import openpyxl
import requests
import os
from datetime import datetime

# Configurações
EXCEL_PATH = os.path.join(os.path.dirname(__file__), 'base_credenciamentos_2026.xlsx')
API_URL = 'https://darkgoldenrod-pelican-495804.hostingersite.com/sync.php'
API_KEY = 'painel_cred_2026_key'

def parse_date(val):
    if isinstance(val, datetime):
        if val.year < 1900:
            return None
        return val.strftime('%Y-%m-%d')
    if isinstance(val, str) and val.strip():
        try:
            return datetime.strptime(val.strip(), '%Y-%m-%d').strftime('%Y-%m-%d')
        except:
            return None
    return None

def to_int(val):
    try:
        return int(val) if val is not None else 0
    except:
        return 0

def to_str(val):
    if val is None:
        return ''
    return str(val).strip()

def load_excel():
    print(f'Lendo {EXCEL_PATH}...')
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb['Base']

    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    print(f'Colunas encontradas: {headers}')

    records = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in row):
            continue
        r = dict(zip(headers, row))
        records.append({
            'uf':                     to_str(r.get('UF')),
            'ibge':                   to_int(r.get('IBGE')),
            'ibge_validacao':         to_int(r.get('IBGE_Validação')),
            'uf_validacao':           to_str(r.get('UF_validação')),
            'municipio':              to_str(r.get('Município')),
            'regiao':                 to_str(r.get('Região')),
            'macrorregiao':           to_str(r.get('Macrorregião')),
            'regiao_saude':           to_str(r.get('Região de saúde')),
            'ied':                    to_int(r.get('IED')),
            'tipo':                   to_str(r.get('Tipo')),
            'estrategia':             to_str(r.get('Estratégia')),
            'nome_completo':          to_str(r.get('Nome completo')),
            'portaria':               to_int(r.get('Portaria')),
            'data_portaria':          parse_date(r.get('Data')),
            'credenciado':            to_int(r.get('Credenciado')),
            'homologacao_1':          to_int(r.get('Homologação 1')),
            'homologacao_2':          to_int(r.get('Homologação 2')),
            'homologacao_3':          to_int(r.get('Homologação 3')),
            'impacto_1':              to_int(r.get('Impacto 1')),
            'impacto_2':              to_int(r.get('Impacto 2')),
            'selecao_credenciamento': to_str(r.get('Seleção Credenciamento')),
            'selecao_homologacao':    to_str(r.get('Seleção homologação')),
            'observacao':             to_str(r.get('Observação')),
            'mensagem_homologacao':   to_str(r.get('Mensagem Homologação')),
            'mensagem_painel':        to_str(r.get('Mensagem painel')),
            'ano':                    to_int(r.get('Ano')),
        })

    print(f'{len(records)} registros carregados.')
    return records

def sync(records):
    print('Enviando para o servidor...')
    # Envia em lotes de 1000 para não estourar o limite
    batch_size = 1000
    total = 0

    # Primeiro envio: limpa e insere o primeiro lote
    first_batch = records[:batch_size]
    resp = requests.post(
        f'{API_URL}?action=sync',
        json={'records': first_batch},
        headers={'X-Api-Key': API_KEY},
        timeout=60
    )
    resp.raise_for_status()
    result = resp.json()
    total += result.get('inserted', 0)
    print(f'Lote 1: {total} registros inseridos')

    # Lotes seguintes: apenas insere (sem truncar)
    for i in range(batch_size, len(records), batch_size):
        batch = records[i:i+batch_size]
        resp = requests.post(
            f'{API_URL}?action=append',
            json={'records': batch},
            headers={'X-Api-Key': API_KEY},
            timeout=60
        )
        resp.raise_for_status()
        result = resp.json()
        total += result.get('inserted', 0)
        print(f'Lote {i//batch_size + 1}: {total} registros inseridos no total')

    print(f'Sincronização concluída! {total} registros no banco.')

if __name__ == '__main__':
    records = load_excel()
    sync(records)
